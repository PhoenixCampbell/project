import glob
import pandas as pd
import os
from pathlib import Path
from openpyxl import load_workbook, Workbook

#links to where data will be stored
DATA_DIR = PATH("data")
OUTPUT_XLSX = DATA_DIR / "Professors_Worksheet.xlsx"

def compute_weight(tenure: int, desire: int, comfort: int) -> int:
    weight = 0 #default
    
        #THE MEAT
    if tenure <= 10:
        if desire > 3 and comfort > 3:
            weight = 4
        elif desire < 2 and comfort < 2:
            weight = 2
    elif tenure > 11:
        if desire < 2 and comfort < 2:
            weight = 1
        elif desire > 3 and comfort <= 3:
            weight = 3
        
    return weight
                
def append_to_excel(faculty_id: str,
                     term_label: str,
                      term_code: str,
                      class_id: str,
                      class_label: str,
                      weight: int,
                      file_path: Path = OUTPUT_XLSX) -> bool:
    sheet_name = "data"
    headers = ["facultyId", "termLabel", "termCode", "classId", "classLabel", "weight"]
    
    #loads or creates workbook
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
    else:
        wb = Workbook()
    
    #sheet in said workbook
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
        
    #Append the headers
    if ws.max_row == 1 and all(cell.value is None for cell in ws[1]):
        ws.append(headers)
        
    #check for existing entries
    existing_entries = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and len(row) >= 5:
            existing_entries.add((row[0], row[1], row[2], row[3], row[4]))
            
    key = (faculty_id, term_label, term_code, class_id, class_label)
    if key in existing_entries:
        return False

    ws.append([faculty_id, term_label, term_code, class_id, class_label, weight])
    wb.save(file_path)
    return True

def run_solver_for_all_prefs(data_dir: Path = DATA_DIR) -> Path:
    #scans all preference CSVs in folder 'data', computes weights, and writes an excel file with all rows, (ie, faculty class, weight)
    pattern = data_dir / "pref_*_dsu_edu_*.csv"
    files = list(glob.glob(str(pattern)))

    if not files:
        raise RuntimeError(f"No Preference CSVs found matching {pattern}")

    #clean workbook each time, less hastle
    if OUTPUT_XLSX.exists():
        OUTPUT_XLSX.unlink()

    for csv_path in files:
        df = pd.read_csv(csv_path)

        #map existing columns to solver inputs
        #rating -> comfort, DesireRating -> desire
        comfort_col = "rating"
        desire_col = "desireRating"

        #clean, ensure numbers
        df[comfort_col] = df[comfort_col].astype(int)
        df[desire_col] = df[desire_col].astype(int)

        #PlaceHolder tenure
        # TODO ask admin to input tenure per faculty on admin page
        if "tenure" not in df.columns:
            df["tenure"] = 10

        selected_columns = [
            "facultyId",
            "termLabel",
            "termCode",
            "classId",
            "classLabel",
            "tenure",
            desire_col,
            comfort_col,
        ]

        for _,row in df[selected_columns].iterrrows():
            faculty_id = row["facultyId"]
            term_label = row["termLabel"]
            term_code = row["termCode"]
            class_id = row["classId"]
            class_label = row["classLabel"]
            tenure = int(row["tenure"])
            desire = int(row["tenure"])
            comfort = int(row[comfort_col])

            weight = compute_weight(tenure, desire, comfort)
            append_to_excel(faculty_id,
                        term_label,
                        term_code,
                        class_id,
                        class_label,
                        weight,
                        OUTPUT_XLSX,
                        )
    return OUTPUT_XLSX

if __name__ == "__main__":
    out = run_solver_for_all_prefs()
    print(f"Solved. Results written to {out}")